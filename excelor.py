from ctypes.wintypes import tagRECT
from distutils import extension
from fileinput import filename
from hashlib import new
from msilib.schema import Patch
from sys import float_repr_style
from turtle import pd
from openpyxl import Workbook
#from openpyxl.utils import get_column_letterfrom
from openpyxl import load_workbook
import pandas as pd
import glob, os, shutil, sys
from PIL import Image
from importlib.metadata import files
from logging import lastResort
from pathlib import Path
from datetime import datetime

# Cargar el archivo xls más reciente

folder = Path(os.getcwd())

def basuraitor():
    for archivo_ruta in os.scandir(folder):
        if archivo_ruta.is_file() and archivo_ruta.name.endswith('.csv'):
           # nombre archivo
           print('este es el archivo ruta', archivo_ruta)
           origen_archivo = archivo_ruta
           # ruta completa
           print('esta es archivo ruta path', archivo_ruta.path)
           # nombre del archivo 
           print('este es el archivo_ruta.name', archivo_ruta.name)
           new_dir = archivo_ruta.name.rsplit('.', 1)[0]
           print('Este es el archivo', new_dir)
           #print(new_dir)
        try:
           # Crea el directorio uniendo ruta y nombre del archivo
           fechar = (datetime.today().strftime('_''%Y_%m_%d_%H_%M'))
           print(fechar)
           ruta_nueva_NoDate = str(Path(os.path.join(folder, new_dir)))
           print(ruta_nueva_NoDate)
           ruta_nueva = ruta_nueva_NoDate + fechar
           print(ruta_nueva)
           os.makedirs(ruta_nueva)
        except WindowsError:
        # Ventana de error si el directorio existe
        # Handle the case where the target dir already exist.
            pass
    #Mueve el archivo al directorio creado
            directorio_nuevo = Path(ruta_nueva)
            print('esta es directorio nuevo',directorio_nuevo)
            shutil.move(Path(origen_archivo), Path(directorio_nuevo))

basuraitor()













# Cargar el archivo y imprime hojas
'''

wb = load_workbook(filename = "Muebles_dico_pa_Abril_22.xlsx")
print(wb.sheetnames)
# elegir hoja
wb.active = wb['Soporte Técnico Avanzado']
#
hoja = wb.active
#print(hoja)
celda = hoja['C37']
celda1 = hoja['C38']

# 

"""
if celda.value == celda1.value:
  print("Value of the Cell 1:",celda.value)
  print("ahuevirri")
else: 
  print("seas mamon") 
"""

valor = 2700

if celda.value == valor:
  print("Value of the Cell 1:",celda.value)
  print("ahuevirri")
else: 
  print("seas mamon") 

'''
